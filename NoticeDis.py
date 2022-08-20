import discord
import openpyxl
import datetime
from discord.ext import commands
import string

client = discord.Client()
cclient = commands.Bot()
permission = discord.Permissions()

SUG_channel = 616922772295122955
TOKEN = ""
COLOR = 0x000000
filename = "공지.xlsx"
now = datetime.datetime.now()
date = str(now.year) + ". " + str(now.month) + ". " + str(now.day)
time = str(now.hour) + ":" + str(now.minute)


async def print_embed(message, text):
    embed = discord.Embed(color=0x000000)
    embed.set_footer(text=text)
    await message.channel.send(embed=embed)


def user_tag(name):
    file = openpyxl.load_workbook("유저목록.xlsx")
    sheet = file.active
    search_isn_data_anfi(name, "유저목록.xlsx")


def search_is_data(data):
    file = openpyxl.load_workbook(filename)
    sheet = file.active
    i = 1
    while True:
        if sheet["A" + str(i)].value != data:
            if sheet["A" + str(i)].value == None:
                return i
            i += 1
        if sheet["A" + str(i)].value == data:
            return 0


def search_isn_data(data):
    file = openpyxl.load_workbook(filename)
    sheet = file.active
    i = 1
    while True:
        if sheet["A" + str(i)].value != data:
            if sheet["A" + str(i)].value == None:
                return 0
            i += 1
        if sheet["A" + str(i)].value == data:
            return i


def search_isn_data_anfi(data, filename):
    file = openpyxl.load_workbook(filename)
    sheet = file.active
    i = 1
    while True:
        if sheet["A" + str(i)].value != data:
            if sheet["A" + str(i)].value == None:
                return 0
            i += 1
        if sheet["A" + str(i)].value == data:
            return i


async def notice_list(message):
    file = openpyxl.load_workbook(filename)
    sheet = file.active
    embed = discord.Embed(color=COLOR)
    i = 1
    while True:
        if sheet["A" + str(i)].value != None:
            if sheet["B"+str(i)].value != None:
                embed.add_field(name=sheet["A"+str(i)].value, value=sheet["B"+str(i)].value[:10], inline=False)
                i += 1
            else:
                embed.add_field(name=sheet["A" + str(i)].value, value="내용없음",
                                inline=False)
                i += 1
        if sheet["A" + str(i)].value == None:
            break
    await message.channel.send(embed=embed)
    for j in range(0, i):
        embed.remove_field(0)


async def input_notice(message, name="NAME"):
    file = openpyxl.load_workbook(filename)
    sheet = file.active
    sid = search_is_data(name)
    if sid == 0:
        await print_embed(message, "해당 공지가 이미 존재합니다.")
        file.save(filename)

    if sid != 0:
        sheet["A" + str(sid)].value = name
        await print_embed(message, "\""+name+"\"공지가 생성되었습니다")
        file.save(filename)


async def delete_notice(message, name):
    file = openpyxl.load_workbook(filename)
    sheet = file.active
    sind = search_isn_data(name)
    if sind == 0:
        await print_embed(message, "해당 공지는 존재하지 않습니다.")
    else:
        sheet.delete_rows(sind)
        file.save(filename)
        await print_embed("\""+name+"\" 공지를 삭제하였습니다.")


async def set_title(message, name, title):
    file = openpyxl.load_workbook(filename)
    sheet = file.active
    sind = search_isn_data(name)
    if sind == 0:
        await print_embed(message, "해당 공지는 존재하지 않습니다.")
        file.save(filename)
    else:
        sheet["B" + str(sind)].value = title
        await print_embed(message, "\""+name+"\" 공지의 제목을 \""+title+"\" 로 설정하였습니다.")
        file.save(filename)


async def set_message(message, name, in_message):
    file = openpyxl.load_workbook(filename)
    sheet = file.active
    sind = search_isn_data(name)
    if sind == 0:
        await print_embed(message, "해당 공지는 존재하지 않습니다.")
        file.save(filename)
    else:
        in_message.replace('~', '\n')
        sheet["C" + str(sind)].value = in_message
        await print_embed(message, "\""+name+"\" 공지의 메시지를 설정하였습니다.")
        file.save(filename)


async def set_color(message, name, color):
    file = openpyxl.load_workbook(filename)
    sheet = file.active
    sind = search_isn_data(name)
    if sind == 0:
        await print_embed(message, "해당 공지는 존재하지 않습니다.")
        file.save(filename)
    else:
        sheet["D" + str(sind)].value = color
        await print_embed(message, "\""+name+"\" 공지의 컬러를 \""+color+"\" 로 설정하였습니다.")
        file.save(filename)


async def send_notice(message, name):
    file = openpyxl.load_workbook(filename)
    sheet = file.active
    sind = search_isn_data(name)
    if sind == 0:
        await print_embed(message, "해당 공지는 존재하지 않습니다.")
        file.save(filename)
    else:
        embed = discord.Embed()
        color = sheet["D" + str(sind)].value
        if sheet["B" + str(sind)].value == None:
            await print_embed(message, "해당 공지의 제목이 설정되지 않았습니다.")

        out_message = sheet["C" + str(sind)].value
        if out_message == None:
            await print_embed(message, "해당 공지의 메시지가 설정되지 않았습니다.")
        if color == None:
            await print_embed(message, "해당 공지의 색상이 설정되지 않았습니다.")
        if color.lower() == "red":
            embed = discord.Embed(color=0xff0000)
        if color.lower() == "green":
            embed = discord.Embed(color=0x00ff00)
        if color.lower() == "blue":
            embed = discord.Embed(color=0x0000ff)
        table = str.maketrans('~', '\n')

        text = str(client.get_user(message.author.id)) + " | " + date + "  •  " + time
        embed.add_field(name=sheet["B" + str(sind)].value, value=out_message.translate(table))
        embed.set_footer(icon_url=message.author.avatar_url, text=text)
        await message.channel.send("[@everyone]")
        await message.channel.send(embed=embed)
        await message.delete()
        file.save(filename)


async def send_text(message, color, text):
    if color.lower() == "red":
        color = 0xff0000
    elif color.lower() == "green":
        color = 0x00ff00
    elif color.lower() == "blue":
        color = 0x0000ff
    embed = discord.Embed(color=int(color))
    table = str.maketrans('~', '\n')
    text = text.translate(table)
    name = str(client.get_user(message.author.id))
    embed.set_author(icon_url=message.author.avatar_url, name=name)
    embed.add_field(name="-", value=text)

    await message.delete()
    await message.channel.send(embed=embed)


async def suggest(message):
    embed = discord.Embed(color=0xFACC2E)
    name = str(client.get_user(message.author.id))
    text = name + " | " + date + "  •  " + time
    embed.set_footer(icon_url=message.author.avatar_url, text=text)
    table = str.maketrans('~', '\n')
    text = message.content.translate(table)
    embed.add_field(name="#건의사항", value=text)
    await message.channel.send(embed=embed)
    await message.delete()


@client.event
async def on_ready():
    game = discord.Game("Minecraft")
    print("login")
    print(client.user.name)
    print(client.user.id)
    print("------------------")
    await client.change_presence(status=discord.Status.idle, activity=game)


@client.event
async def on_message(message):
    if message.channel.id == SUG_channel:
        await suggest(message)
    else:
        if message.content.startswith("/공지"):
            file = openpyxl.load_workbook(filename)
            sheet = file.active
            arg = message.content.split(" ")
            if arg[1] == "목록":
                await notice_list(message)
            if arg[1] == "생성":
                await input_notice(message, arg[2])
            if arg[1] == "삭제":
                await delete_notice(message, arg[2])
            if arg[1] == "제목":
                await set_title(message, arg[2], arg[3])
            if arg[1] == "메시지":
                await set_message(message, arg[2], message.content[9+len(arg[2]):])
            if arg[1] == "색상":
                await set_color(message, arg[2], arg[3])
            if arg[1] == "전송":
                await send_notice(message, arg[2])
        if message.content.startswith("/텍스트"):
            arg = message.content.split(" ")
            await send_text(message, arg[1], message.content[6+len(arg[1]):])


client.run(TOKEN)
